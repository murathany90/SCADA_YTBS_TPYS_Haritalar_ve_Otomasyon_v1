#!/usr/bin/env python3
"""
Builds the V2 YTBS/KML model by enriching the current KML source with
TM / hat / trafo / bara Excel lists and SCADA mapping rows.
"""

from __future__ import annotations

import json
import math
import re
import unicodedata
import warnings
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl


ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "docs" / "yeni_harita_modeli"
OUTPUT_JSON = ROOT / "data" / "kml_layers_v2.json"
OUTPUT_REPORT = INPUT_DIR / "kml_layers_v2_validation.md"

NS = {"kml": "http://www.opengis.net/kml/2.2"}
SCHEMA_VERSION = 2
DEFAULT_YTM = "Orta Anadolu YTM"
VALID_BARAGE_SCADA_KV = {"154", "400"}

TM_FILE = INPUT_DIR / "01-TRAFO_MERKEZI_LISTESI.xlsx"
BARA_FILE = INPUT_DIR / "02-BARA_LISTESI.xlsx"
HAT_FILE = INPUT_DIR / "09-HAT_LISTESI.xlsx"
TRAFO_FILE = INPUT_DIR / "11-TRAFO_LISTESI.xlsx"
KML_FILE = INPUT_DIR / "20-YTBS_Detayli_Harita (3).kml"
SCADA_FILE = INPUT_DIR / "SISTEM_ESLEME_LISTESI.xlsx"
MATCH_TABLE_FILE = INPUT_DIR / "eslesme_tablolari.xlsx"


def normalize_text(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    raw = raw.replace("İ", "I").replace("ı", "i")
    raw = raw.replace("Ş", "S").replace("ş", "s")
    raw = raw.replace("Ğ", "G").replace("ğ", "g")
    raw = raw.replace("Ü", "U").replace("ü", "u")
    raw = raw.replace("Ö", "O").replace("ö", "o")
    raw = raw.replace("Ç", "C").replace("ç", "c")
    raw = re.sub(r"\s+", " ", raw)
    return raw.upper().strip()


def clean_string(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text)


def number_to_string(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value):
            if value.is_integer():
                return str(int(value))
            return f"{value:.12g}"
        return ""
    return clean_string(value)


def as_int_string(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return clean_string(value)
    return str(int(number)) if math.isfinite(number) else ""


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def excel_date_to_iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if value is None or value == "":
        return None
    text = clean_string(value)
    return text or None


def sort_key(value: Any) -> tuple[int, Any]:
    text = number_to_string(value)
    if re.fullmatch(r"-?\d+", text):
        return (0, int(text))
    return (1, normalize_text(text))


def load_excel_rows(path: Path) -> list[dict[str, Any]]:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(min_row=1, values_only=True)
    headers = [str(cell) if cell is not None else "" for cell in next(iterator)]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not values or not isinstance(values[0], (int, float)):
            continue
        row = {}
        for idx, header in enumerate(headers):
            row[header] = values[idx] if idx < len(values) else None
        if "AdÄ±" in row and "Adı" not in row:
            row["Adı"] = row["AdÄ±"]
        rows.append(row)
    return rows


def load_table_rows(path: Path, sheet_name_part: str) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_key = normalize_text(sheet_name_part)
    worksheet = next(
        (workbook[name] for name in workbook.sheetnames if sheet_key in normalize_text(name)),
        workbook[workbook.sheetnames[0]],
    )
    iterator = worksheet.iter_rows(min_row=1, values_only=True)
    headers = [str(cell) if cell is not None else "" for cell in next(iterator)]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not values or all(value is None or value == "" for value in values):
            continue
        row = {}
        for idx, header in enumerate(headers):
            row[header] = values[idx] if idx < len(values) else None
        rows.append(row)
    return rows


def get_by_normalized_key(row: dict[str, Any], *keys: str) -> Any:
    normalized = {normalize_text(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(normalize_text(key))
        if value is not None:
            return value
    return None


def get_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def parse_coordinate_sequence(text: str | None) -> list[list[float]]:
    if not text:
        return []
    coords: list[list[float]] = []
    for token in text.replace("\n", " ").replace("\t", " ").split():
        parts = token.split(",")
        if len(parts) < 2:
            continue
        lon = as_float(parts[0])
        lat = as_float(parts[1])
        if lon is None or lat is None:
            continue
        coords.append([round(lon, 6), round(lat, 6)])
    return coords


def parse_bbox(coords: list[list[float]]) -> list[float] | None:
    if not coords:
        return None
    lons = [coord[0] for coord in coords]
    lats = [coord[1] for coord in coords]
    return [round(min(lons), 6), round(min(lats), 6), round(max(lons), 6), round(max(lats), 6)]


def infer_kv(*values: Any) -> str:
    for value in values:
        text = clean_string(value)
        if not text:
            continue
        match = re.search(r"(400|380|154|66)\s*kV", text, re.IGNORECASE)
        if match:
            return match.group(1)
        match = re.search(r"(400|380|154|66)\s*KV", normalize_text(text))
        if match:
            return match.group(1)
    for value in values:
        text = clean_string(value)
        if not text:
            continue
        match = re.search(r"(400|380|154|66)", text)
        if match:
            return match.group(1)
    return ""


def walk_kml_tree(node: ET.Element, path_parts: list[str], out: list[dict[str, Any]]) -> None:
    children = list(node)
    for child in children:
        local_name = get_local_name(child.tag)
        if local_name == "Folder":
            folder_name = clean_string(child.findtext("kml:name", default="", namespaces=NS))
            next_path = path_parts + ([folder_name] if folder_name else [])
            walk_kml_tree(child, next_path, out)
            continue
        if local_name != "Placemark":
            walk_kml_tree(child, path_parts, out)
            continue

        name = clean_string(child.findtext("kml:name", default="", namespaces=NS))
        description = clean_string(child.findtext("kml:description", default="", namespaces=NS))
        style_url = clean_string(child.findtext("kml:styleUrl", default="", namespaces=NS))

        point = child.find(".//kml:Point/kml:coordinates", NS)
        line_strings = child.findall(".//kml:LineString/kml:coordinates", NS)

        if point is not None:
            coords = parse_coordinate_sequence(point.text)
            if coords:
                out.append(
                    {
                        "kind": "tm",
                        "name": name,
                        "kmlDescriptionId": description,
                        "styleUrl": style_url,
                        "coords": coords,
                        "folder": " / ".join(path_parts),
                        "kv": infer_kv(style_url, name),
                    }
                )
            continue

        if line_strings:
            coords: list[list[float]] = []
            for coord_node in line_strings:
                seq = parse_coordinate_sequence(coord_node.text)
                if not seq:
                    continue
                if coords and coords[-1] == seq[0]:
                    coords.extend(seq[1:])
                else:
                    coords.extend(seq)
            if coords:
                out.append(
                    {
                        "kind": "hat",
                        "name": name,
                        "kmlDescriptionId": description,
                        "styleUrl": style_url,
                        "coords": coords,
                        "bbox": parse_bbox(coords),
                        "folder": " / ".join(path_parts),
                        "kv": infer_kv(style_url, name),
                    }
                )


def load_kml_features(path: Path) -> dict[str, list[dict[str, Any]]]:
    tree = ET.parse(path)
    root = tree.getroot()
    features: list[dict[str, Any]] = []
    walk_kml_tree(root, [], features)
    tm_features = [feature for feature in features if feature["kind"] == "tm"]
    hat_features = [feature for feature in features if feature["kind"] == "hat"]
    return {"tm": tm_features, "hat": hat_features}


def empty_scada_metric() -> dict[str, Any]:
    return {
        "ids": [],
        "rows": [],
        "resolvedId": None,
        "ambiguous": False,
    }


def new_scada_group(include_voltage: bool = False) -> dict[str, Any]:
    group = {
        "active": empty_scada_metric(),
        "reactive": empty_scada_metric(),
    }
    if include_voltage:
        group["voltage"] = empty_scada_metric()
    return group


def parse_formula_parts(formula_raw: Any) -> list[dict[str, Any]]:
    text = str(formula_raw or "")
    parts: list[dict[str, Any]] = []
    pattern = re.compile(r"^\(([-+]?\d+)\)\s*([^,]+),\s*([^,]+),\s*([^,]+),\s*(.+?)$")
    for line in text.splitlines():
        raw = clean_string(line)
        if not raw:
            continue
        match = pattern.match(raw)
        if not match:
            parts.append({"raw": raw, "parsed": False})
            continue
        parts.append(
            {
                "raw": raw,
                "parsed": True,
                "sign": int(match.group(1)),
                "stationCode": clean_string(match.group(2)),
                "kv": clean_string(match.group(3)),
                "targetCode": clean_string(match.group(4)),
                "quantity": clean_string(match.group(5)),
            }
        )
    return parts


def make_scada_candidate(row: dict[str, Any]) -> dict[str, Any]:
    formula_raw = clean_string(row["ÖLÇÜM NOKTASI FORMÜLASYONU"])
    formula_parts = parse_formula_parts(formula_raw)
    primary_part = next((part for part in formula_parts if part.get("parsed")), None)
    return {
        "sourceId": as_int_string(row["ID"]),
        "analogId": as_int_string(row["ANALOG ÖLÇÜM ID"]),
        "measurementId": clean_string(row["ÖLÇÜM NOKTASI ID"]),
        "analogName": clean_string(row["ANALOG ÖLÇÜM"]),
        "formulaRaw": formula_raw,
        "formulaParts": formula_parts,
        "formulaSign": primary_part.get("sign") if primary_part else None,
        "formulaStationCode": primary_part.get("stationCode") if primary_part else "",
        "formulaTargetCode": primary_part.get("targetCode") if primary_part else "",
    }


def build_terminal_ref(name: Any, tm_entity: dict[str, Any] | None) -> dict[str, str]:
    return {
        "name": clean_string(tm_entity["name"]) if tm_entity else clean_string(name),
        "ucteKodu": clean_string(tm_entity["ucteKodu"]) if tm_entity else "",
        "psseAdi": clean_string(tm_entity["psseAdi"]) if tm_entity else "",
    }


def resolve_terminal_side(
    value: Any,
    start_ref: dict[str, str],
    end_ref: dict[str, str],
) -> tuple[str, str]:
    raw = clean_string(value)
    if not raw:
        return "unknown", ""

    raw_upper = raw.upper()
    exact_checks = [
        ("tm-name-exact", clean_string(start_ref.get("name")).upper(), clean_string(end_ref.get("name")).upper()),
        ("ucte-exact", clean_string(start_ref.get("ucteKodu")).upper(), clean_string(end_ref.get("ucteKodu")).upper()),
        ("psse-exact", clean_string(start_ref.get("psseAdi")).upper(), clean_string(end_ref.get("psseAdi")).upper()),
    ]
    for basis, start_value, end_value in exact_checks:
        start_match = bool(start_value) and raw_upper == start_value
        end_match = bool(end_value) and raw_upper == end_value
        if start_match and not end_match:
            return "start", basis
        if end_match and not start_match:
            return "end", basis
        if start_match and end_match:
            return "unknown", f"{basis}-conflict"

    normalized = normalize_text(raw)
    if not normalized:
        return "unknown", ""

    start_aliases = {
        normalize_text(start_ref.get("name")),
        normalize_text(start_ref.get("ucteKodu")),
        normalize_text(start_ref.get("psseAdi")),
    }
    end_aliases = {
        normalize_text(end_ref.get("name")),
        normalize_text(end_ref.get("ucteKodu")),
        normalize_text(end_ref.get("psseAdi")),
    }
    start_aliases.discard("")
    end_aliases.discard("")
    start_match = normalized in start_aliases
    end_match = normalized in end_aliases
    if start_match and not end_match:
        return "start", "normalized-alias-exact"
    if end_match and not start_match:
        return "end", "normalized-alias-exact"
    if start_match and end_match:
        return "unknown", "normalized-alias-conflict"
    return "unknown", ""


def enrich_hat_candidate(
    candidate: dict[str, Any],
    row: dict[str, Any],
    hat: dict[str, Any],
    tm_by_id: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    enriched = dict(candidate)
    source_tm_name = clean_string(row["TRAFO MERKEZİ"])
    start_ref = build_terminal_ref(hat.get("startTm"), tm_by_id.get(str(hat.get("startTmId") or "")))
    end_ref = build_terminal_ref(hat.get("endTm"), tm_by_id.get(str(hat.get("endTmId") or "")))

    source_side, source_basis = resolve_terminal_side(source_tm_name, start_ref, end_ref)
    station_side, station_basis = resolve_terminal_side(enriched.get("formulaStationCode"), start_ref, end_ref)
    target_side, target_basis = resolve_terminal_side(enriched.get("formulaTargetCode"), start_ref, end_ref)

    terminal_side = source_side
    terminal_match_basis = source_basis
    if terminal_side == "unknown" and station_side != "unknown":
        terminal_side = station_side
        terminal_match_basis = station_basis or "formula-station"
    if terminal_side == "unknown" and target_side == "end":
        terminal_side = "start"
        terminal_match_basis = f"opposite-{target_basis or 'formula-target'}"
    elif terminal_side == "unknown" and target_side == "start":
        terminal_side = "end"
        terminal_match_basis = f"opposite-{target_basis or 'formula-target'}"

    polarization_sign = 1 if terminal_side == "start" else -1 if terminal_side == "end" else None
    formula_sign = enriched.get("formulaSign")
    polarization_consistent = None
    if polarization_sign is not None and formula_sign in (1, -1):
        polarization_consistent = formula_sign == polarization_sign

    if target_side == "unknown" and terminal_side == "start":
        inferred_target_side = "end"
    elif target_side == "unknown" and terminal_side == "end":
        inferred_target_side = "start"
    else:
        inferred_target_side = target_side

    enriched.update(
        {
            "sourceTmName": source_tm_name,
            "sourceSide": terminal_side,
            "targetSide": inferred_target_side,
            "candidateSlot": "primary" if terminal_side == "start" else "secondary" if terminal_side == "end" else "extra",
            "terminalSide": terminal_side,
            "terminalMatchBasis": terminal_match_basis,
            "polarizationSign": polarization_sign,
            "polarizationConsistent": polarization_consistent,
            "sourceTmNormalized": normalize_text(source_tm_name),
            "startTmNormalized": normalize_text(start_ref.get("name")),
            "endTmNormalized": normalize_text(end_ref.get("name")),
        }
    )
    return enriched


def append_scada_metric(metric: dict[str, Any], candidate: dict[str, Any]) -> None:
    measurement_id = clean_string(candidate.get("measurementId"))
    existing_row = next(
        (
            row
            for row in metric["rows"]
            if clean_string(row.get("measurementId")) == measurement_id and measurement_id
        ),
        None,
    )
    if existing_row:
        existing_variants = {
            clean_string(existing_row.get("formulaRaw")),
            *(clean_string(item) for item in existing_row.get("formulaVariants") or []),
        }
        candidate_formula = clean_string(candidate.get("formulaRaw"))
        if candidate_formula:
            existing_variants.add(candidate_formula)
        existing_row["formulaVariants"] = sorted(item for item in existing_variants if item)
        if (
            candidate_formula
            and candidate_formula != clean_string(existing_row.get("formulaRaw"))
        ) or (
            existing_row.get("formulaSign") != candidate.get("formulaSign")
        ) or (
            clean_string(existing_row.get("terminalSide")) != clean_string(candidate.get("terminalSide"))
        ):
            metric["ambiguous"] = True
            metric["singleIdDualFormulaConflict"] = True
        return

    metric["rows"].append(candidate)
    if measurement_id and measurement_id not in metric["ids"]:
        metric["ids"].append(measurement_id)
    if (
        len(metric["rows"]) > 1
        or len(metric["ids"]) > 1
        or len(candidate.get("formulaParts") or []) > 1
        or any(not part.get("parsed", False) for part in candidate.get("formulaParts") or [])
    ):
        metric["ambiguous"] = True


def append_scada_metric_once(metric: dict[str, Any], candidate: dict[str, Any]) -> bool:
    measurement_id = clean_string(candidate.get("measurementId"))
    if not measurement_id or measurement_id in metric.get("ids", []):
        return False
    append_scada_metric(metric, candidate)
    return True


def parse_measure_kind(analog_text: str) -> tuple[str, str]:
    parts = analog_text.split(",", 1)
    kind = clean_string(parts[0]) if parts else ""
    detail = clean_string(parts[1]) if len(parts) > 1 else ""
    return kind, detail


def metric_key_from_kind(kind: str) -> str | None:
    normalized = normalize_text(kind)
    if normalized == "AKTIF GUC (MW)":
        return "active"
    if normalized == "REAKTIF GUC (MVAR)":
        return "reactive"
    if normalized == "GERILIM (KV)":
        return "voltage"
    return None


def compact_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize_text(value))


def voltage_bucket(value: Any) -> str:
    text = number_to_string(value) or clean_string(value)
    match = re.search(r"(400|380|154|66)", text)
    if not match:
        return ""
    bucket = match.group(1)
    return "400" if bucket == "380" else bucket


def bara_alias_variants(*values: Any) -> set[str]:
    variants: set[str] = set()
    for value in values:
        compact = compact_key(value)
        if not compact:
            continue
        candidates = {compact}
        without_kv = re.sub(r"^(400|380|154|66)", "", compact)
        if without_kv:
            candidates.add(without_kv)
        expanded: set[str] = set()
        for candidate in candidates:
            expanded.add(candidate)
            if candidate.startswith("BB") and len(candidate) > 2:
                expanded.add(f"B{candidate[2:]}")
            if candidate.startswith("BARA") and len(candidate) > 4:
                expanded.add(f"B{candidate[4:]}")
        variants.update(alias for alias in expanded if alias and len(alias) >= 2)
    return variants


def make_voltage_overlay_candidate(row: dict[str, Any], source: str) -> dict[str, Any]:
    measurement_id = clean_string(get_by_normalized_key(row, "Ölçüm Noktası ID", "Olcum Noktasi ID"))
    formula_body = clean_string(get_by_normalized_key(row, "Ölçüm Noktası Formülasyonu (Katsayısız)", "Olcum Noktasi Formulasyonu (Katsayisiz)"))
    sign_value = as_int_string(get_by_normalized_key(row, "Ölçüm Noktası Formülasyon Katsayısı", "Olcum Noktasi Formulasyon Katsayisi")) or "1"
    sign = -1 if str(sign_value).strip().startswith("-") else 1
    formula_raw = f"({'+' if sign >= 0 else '-'}1) {formula_body}" if formula_body else ""
    formula_parts = parse_formula_parts(formula_raw)
    primary_part = next((part for part in formula_parts if part.get("parsed")), None)
    return {
        "sourceId": "",
        "analogId": "",
        "measurementId": measurement_id,
        "analogName": f"Gerilim (kV), {clean_string(get_by_normalized_key(row, 'Bara Adı', 'Bara Adi'))}",
        "formulaRaw": formula_raw,
        "formulaParts": formula_parts,
        "formulaSign": primary_part.get("sign") if primary_part else sign,
        "formulaStationCode": primary_part.get("stationCode") if primary_part else "",
        "formulaTargetCode": primary_part.get("targetCode") if primary_part else "",
        "source": source,
    }


def new_tm_entity(feature: dict[str, Any], excel_row: dict[str, Any]) -> dict[str, Any]:
    coords = feature["coords"][0]
    return {
        "id": as_int_string(excel_row["ID"]),
        "kmlDescriptionId": as_int_string(feature["kmlDescriptionId"]),
        "name": clean_string(excel_row["Adı"]) or feature["name"],
        "lon": coords[0],
        "lat": coords[1],
        "folder": feature["folder"],
        "kv": feature["kv"],
        "ytm": clean_string(excel_row["Yük Tevzi Müdürlüğü"]),
        "il": clean_string(excel_row["İli"]),
        "bolgeMudurlugu": clean_string(excel_row["Bölge Müdürlüğü"]),
        "dagitimSirketi": clean_string(excel_row["Dağıtım Şirketi"]),
        "mulk": clean_string(excel_row["Mülkiyet"]),
        "psseNo": as_int_string(excel_row["PSSE NO"]),
        "psseAdi": clean_string(excel_row["PSSE ADI"]),
        "ucteKodu": clean_string(excel_row["UCTE Kodu"]),
        "oysId": clean_string(excel_row["OYS ID"]),
        "saltTuru": clean_string(excel_row["ŞALT TÜRÜ"]),
        "insaYili": as_int_string(excel_row["İnşa Yılı"]),
        "rakimM": as_float(excel_row["Rakım (m)"]),
        "status": clean_string(excel_row["Durumu"]),
        "baslangicTarihi": excel_date_to_iso(excel_row["Başlangıç Tarihi"]),
        "bitisTarihi": excel_date_to_iso(excel_row["Bitiş Tarihi"]),
        "childTrafoIds": [],
        "childBaraIds": [],
        "childHatIds": [],
    }


def new_hat_entity(feature: dict[str, Any], excel_row: dict[str, Any], tm_by_name: dict[str, dict[str, Any]]) -> dict[str, Any]:
    start_tm = clean_string(excel_row["Başlangıç Trafo Merkezi"])
    end_tm = clean_string(excel_row["Bitiş Trafo Merkezi"])
    start_tm_entity = tm_by_name.get(normalize_text(start_tm))
    end_tm_entity = tm_by_name.get(normalize_text(end_tm))
    ytm_names = []
    for tm_entity in (start_tm_entity, end_tm_entity):
        if tm_entity and tm_entity["ytm"] and tm_entity["ytm"] not in ytm_names:
            ytm_names.append(tm_entity["ytm"])

    kv = number_to_string(excel_row["Gerilimi (kV)"]) or feature["kv"]
    return {
        "id": as_int_string(excel_row["ID"]),
        "kmlDescriptionId": as_int_string(feature["kmlDescriptionId"]),
        "name": clean_string(excel_row["Adı"]) or feature["name"],
        "coords": feature["coords"],
        "bbox": feature["bbox"],
        "folder": feature["folder"],
        "kv": kv,
        "startTm": start_tm,
        "endTm": end_tm,
        "startTmId": start_tm_entity["id"] if start_tm_entity else None,
        "endTmId": end_tm_entity["id"] if end_tm_entity else None,
        "ytmNames": ytm_names,
        "lengthKm": as_float(excel_row["Toplam Uzunluk (km)"]),
        "characteristic": clean_string(excel_row["Karakteristik"]),
        "winterCapacityMva": as_float(excel_row["Kış Kapasitesi (MVA)"]),
        "summerCapacityMva": as_float(excel_row["Yaz Kapasitesi (MVA)"]),
        "operatingCapacityMva": as_float(excel_row["İşletme Kapasitesi (MVA)"]),
        "status": clean_string(excel_row["Durumu"]),
        "normalIsletmeDurumu": clean_string(excel_row["NORMAL İŞLETME DURUMU\t"]),
        "baslangicTarihi": excel_date_to_iso(excel_row["Başlangıç Tarihi"]),
        "bitisTarihi": excel_date_to_iso(excel_row["Bitiş Tarihi"]),
        "scada": new_scada_group(include_voltage=False),
    }


def new_trafo_entity(row: dict[str, Any], tm_entity: dict[str, Any]) -> dict[str, Any]:
    primary_kv = number_to_string(row["Primer Gerilim (kV)"])
    secondary_kv = number_to_string(row["Sekonder Gerilim (kV)"])
    return {
        "id": as_int_string(row["ID"]),
        "tmId": tm_entity["id"],
        "tmName": tm_entity["name"],
        "ytm": tm_entity["ytm"],
        "name": clean_string(row["Adı"]),
        "gerilimTuru": clean_string(row["Gerilim Türü"]),
        "primaryKv": primary_kv,
        "secondaryKv": secondary_kv,
        "onanMva": as_float(row["ONAN Gücü (MVA)"]),
        "onafMva": as_float(row["ONAF Gücü (MVA)"]),
        "ofafMva": as_float(row["OFAF Gücü (MVA)"]),
        "bazGucuMva": as_float(row["Baz Gücü (MVA)"]),
        "empedansUkPct": as_float(row["Empedansı (%uk)"]),
        "baglantiTuru": clean_string(row["Bağlantı Türü"]),
        "ucSargi": bool(row["Üç Sargı"]) if row["Üç Sargı"] is not None else False,
        "normalIsletmeDurumu": clean_string(row["Normal İşletme Durumu"]),
        "status": clean_string(row["Durumu"]),
        "renderMode": "details-only",
        "baslangicTarihi": excel_date_to_iso(row["Başlangıç Tarihi"]),
        "bitisTarihi": excel_date_to_iso(row["Bitiş Tarihi"]),
        "scada": new_scada_group(include_voltage=True),
    }


def new_bara_entity(row: dict[str, Any], tm_entity: dict[str, Any]) -> dict[str, Any]:
    gerilim_kv = number_to_string(row["Gerilim (kV)"])
    voltage_group = "bara-154-400" if gerilim_kv in VALID_BARAGE_SCADA_KV else "trafo-lv" if normalize_text(clean_string(row["Adı"])).startswith("TR") else "other"
    return {
        "id": as_int_string(row["ID"]),
        "tmId": tm_entity["id"],
        "tmName": tm_entity["name"],
        "ytm": tm_entity["ytm"],
        "name": clean_string(row["Adı"]),
        "gerilimSeviyesi": clean_string(row["Gerilim Seviyesi"]),
        "gerilimKv": gerilim_kv,
        "voltageGroup": voltage_group,
        "kullanim": clean_string(row["Kullanım"]),
        "turu": clean_string(row["Türü"]),
        "veriToplama": clean_string(row["Veri Toplama"]),
        "toprakAyiricisiAdi": clean_string(row["Toprak Ayırıcısı Adı"]),
        "ucteKodu": clean_string(row["UCTE Kodu"]),
        "psseNo": as_int_string(row["PSSE NO"]),
        "status": clean_string(row["Durum"]),
        "renderMode": "details-only",
        "baslangicTarihi": excel_date_to_iso(row["Başlangıç Tarihi"]),
        "bitisTarihi": excel_date_to_iso(row["Bitiş Tarihi"]),
        "scada": {"voltage": empty_scada_metric()},
    }


def build_hierarchy(
    tm_points: list[dict[str, Any]],
    hat_lines: list[dict[str, Any]],
    trafos: list[dict[str, Any]],
    baras: list[dict[str, Any]],
) -> dict[str, Any]:
    ytm_map: dict[str, dict[str, Any]] = {}

    def bucket(ytm_name: str, kv: str) -> dict[str, list[str]]:
        ytm_entry = ytm_map.setdefault(ytm_name, {"gerilim": {}})
        return ytm_entry["gerilim"].setdefault(
            kv or "",
            {"hatIds": [], "tmIds": [], "trafoIds": [], "baraIds": []},
        )

    for tm in tm_points:
        if tm["ytm"]:
            bucket(tm["ytm"], tm["kv"])["tmIds"].append(tm["id"])
    for hat in hat_lines:
        for ytm in hat["ytmNames"]:
            bucket(ytm, hat["kv"])["hatIds"].append(hat["id"])
    for trafo in trafos:
        if trafo["ytm"]:
            bucket(trafo["ytm"], trafo["primaryKv"])["trafoIds"].append(trafo["id"])
    for bara in baras:
        if bara["ytm"]:
            bucket(bara["ytm"], bara["gerilimKv"])["baraIds"].append(bara["id"])

    for ytm_entry in ytm_map.values():
        for kv_entry in ytm_entry["gerilim"].values():
            for key in ("hatIds", "tmIds", "trafoIds", "baraIds"):
                kv_entry[key] = sorted(set(kv_entry[key]), key=sort_key)

    ordered_ytms = {}
    for ytm_name in sorted(ytm_map.keys(), key=normalize_text):
        gerilim = ytm_map[ytm_name]["gerilim"]
        ordered_ytms[ytm_name] = {
            "gerilim": {
                kv: gerilim[kv]
                for kv in sorted(gerilim.keys(), key=sort_key)
            }
        }
    return {"ytm": ordered_ytms}


def count_metric_presence(items: list[dict[str, Any]], metric_path: list[str]) -> int:
    count = 0
    for item in items:
        metric = item
        for key in metric_path:
            metric = metric.get(key) if isinstance(metric, dict) else None
        if isinstance(metric, dict) and metric.get("ids"):
            count += 1
    return count


def count_ambiguous_metrics(items: list[dict[str, Any]], metric_paths: list[list[str]]) -> int:
    count = 0
    for item in items:
        ambiguous = False
        for metric_path in metric_paths:
            metric = item
            for key in metric_path:
                metric = metric.get(key) if isinstance(metric, dict) else None
            if isinstance(metric, dict) and metric.get("ambiguous"):
                ambiguous = True
                break
        if ambiguous:
            count += 1
    return count


def build_validation_report(model: dict[str, Any]) -> str:
    validation = model["meta"]["validation"]
    lines = [
        "# KML Layers V2 Validation",
        "",
        f"Uretim zamani: `{model['meta']['generatedAt']}`",
        "",
        "## Kaynak Ozetleri",
        "",
        f"- KML TM placemark: `{validation['kmlTmCount']}`",
        f"- KML hat placemark: `{validation['kmlHatCount']}`",
        f"- Excel TM satiri: `{validation['excelTmCount']}`",
        f"- Excel hat satiri: `{validation['excelHatCount']}`",
        f"- Excel trafo satiri: `{validation['excelTrafoCount']}`",
        f"- Excel bara satiri: `{validation['excelBaraCount']}`",
        "",
        "## KML / Excel Eslesme",
        "",
        f"- TM eslesmesi: `{validation['tmIdMatches']}/{validation['kmlTmCount']}`",
        f"- Hat eslesmesi: `{validation['hatIdMatches']}/{validation['kmlHatCount']}`",
        "",
        "## Parent Baglama",
        "",
        f"- Trafo -> TM: `{validation['trafoParentMatches']}/{validation['excelTrafoCount']}`",
        f"- Bara -> TM: `{validation['baraParentMatches']}/{validation['excelBaraCount']}`",
        "",
        "## SCADA Kapsami",
        "",
        f"- Hat aktif: `{validation['scadaCoverage']['hatActive']}/{validation['excelHatCount']}`",
        f"- Hat reaktif: `{validation['scadaCoverage']['hatReactive']}/{validation['excelHatCount']}`",
        f"- Trafo aktif: `{validation['scadaCoverage']['trafoActive']}/{validation['excelTrafoCount']}`",
        f"- Trafo reaktif: `{validation['scadaCoverage']['trafoReactive']}/{validation['excelTrafoCount']}`",
        f"- Bara gerilim (tum): `{validation['scadaCoverage']['baraVoltageAll']}/{validation['excelBaraCount']}`",
        f"- 154/400 bara gerilim: `{validation['scadaCoverage']['baraVoltage154400']}/{validation['bara154400Count']}`",
        "",
        "## Hat Terminal Polarizasyonu",
        "",
        f"- Terminal tarafi cozulen aday: `{validation['terminalSideResolved']}`",
        f"- Terminal tarafi bilinmeyen aday: `{validation['terminalSideUnknown']}`",
        f"- Polarizasyon uyumsuz config: `{validation['polarizationMismatch']}`",
        "",
        "## Gerilim Overlay",
        "",
        f"- Exact kaynak eslesmesi: `{validation['voltageExactSourceMatched']}`",
        f"- ChatGPT Bara ID overlay: `{validation['voltageOverlayMatched']}`",
        f"- Guvenli alias fallback: `{validation['voltageAliasMatched']}`",
        f"- Alias ambiguous config: `{validation['voltageAliasAmbiguous']}`",
        f"- 154/400 hala eksik: `{validation['voltageStillMissing']}`",
        "",
        "## Ambiguous Kayitlar",
        "",
        f"- Hat metric ambiguous entity: `{validation['ambiguous']['hatEntities']}`",
        f"- Trafo metric ambiguous entity: `{validation['ambiguous']['trafoEntities']}`",
        f"- Bara metric ambiguous entity: `{validation['ambiguous']['baraEntities']}`",
        "",
        "## KML'de Olmayan Excel Kayitlari",
        "",
    ]

    missing_tms = validation.get("missingTmExcelRows", [])
    if not missing_tms:
        lines.append("- Yok")
    else:
        for row in missing_tms:
            lines.append(f"- `{row['id']}` / {row['name']}")

    lines.extend(
        [
            "",
            "## Notlar",
            "",
            "- `trafos` ve `baraNodes` bu fazda `details-only` tutulur; ayri harita koordinati uretilmez.",
            "- Cok adayli SCADA kayitlari dusurulmez; `ids[]`, `rows[]` ve `ambiguous=true` ile saklanir.",
            "- `TARSUS OSB` KML'de olmadigi icin TM listesinde kalir ama `tmPoints` icine giremez.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    tm_rows = load_excel_rows(TM_FILE)
    bara_rows = load_excel_rows(BARA_FILE)
    hat_rows = load_excel_rows(HAT_FILE)
    trafo_rows = load_excel_rows(TRAFO_FILE)
    scada_rows = load_excel_rows(SCADA_FILE)
    kml_features = load_kml_features(KML_FILE)

    tm_excel_by_id = {as_int_string(row["ID"]): row for row in tm_rows}
    hat_excel_by_id = {as_int_string(row["ID"]): row for row in hat_rows}

    tm_points: list[dict[str, Any]] = []
    missing_tm_excel_rows: list[dict[str, str]] = []
    for feature in kml_features["tm"]:
        tm_row = tm_excel_by_id.get(as_int_string(feature["kmlDescriptionId"]))
        if not tm_row:
            continue
        tm_points.append(new_tm_entity(feature, tm_row))

    kml_tm_ids = {as_int_string(feature["kmlDescriptionId"]) for feature in kml_features["tm"]}
    for row in tm_rows:
        tm_id = as_int_string(row["ID"])
        if tm_id not in kml_tm_ids:
            missing_tm_excel_rows.append({"id": tm_id, "name": clean_string(row["Adı"])})

    tm_by_name = {normalize_text(tm["name"]): tm for tm in tm_points}
    hat_lines = []
    for feature in kml_features["hat"]:
        hat_row = hat_excel_by_id.get(as_int_string(feature["kmlDescriptionId"]))
        if not hat_row:
            continue
        hat_lines.append(new_hat_entity(feature, hat_row, tm_by_name))

    hat_by_name = {normalize_text(hat["name"]): hat for hat in hat_lines}
    hat_by_id = {hat["id"]: hat for hat in hat_lines}
    for hat in hat_lines:
        for tm_id in (hat.get("startTmId"), hat.get("endTmId")):
            if tm_id and tm_id in {tm["id"] for tm in tm_points}:
                tm_by_id = next(tm for tm in tm_points if tm["id"] == tm_id)
                tm_by_id["childHatIds"].append(hat["id"])

    tm_points.sort(key=lambda item: sort_key(item["id"]))
    tm_by_id = {tm["id"]: tm for tm in tm_points}

    trafos: list[dict[str, Any]] = []
    trafo_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for row in trafo_rows:
        tm_entity = tm_by_name.get(normalize_text(row["Trafo Merkezi"]))
        if not tm_entity:
            continue
        trafo = new_trafo_entity(row, tm_entity)
        trafos.append(trafo)
        trafo_by_key[(normalize_text(trafo["tmName"]), normalize_text(trafo["name"]))] = trafo
        tm_entity["childTrafoIds"].append(trafo["id"])

    bara_nodes: list[dict[str, Any]] = []
    bara_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    bara_154_400_count = 0
    for row in bara_rows:
        tm_entity = tm_by_name.get(normalize_text(row["Trafo Merkezi"]))
        if not tm_entity:
            continue
        bara = new_bara_entity(row, tm_entity)
        bara_nodes.append(bara)
        bara_by_key[(normalize_text(bara["tmName"]), normalize_text(bara["name"]))] = bara
        tm_entity["childBaraIds"].append(bara["id"])
        if bara["gerilimKv"] in VALID_BARAGE_SCADA_KV:
            bara_154_400_count += 1

    scada_target_rows = [row for row in scada_rows if normalize_text(row["SİSTEM TÜRÜ"]) == "SCADA"]
    for row in scada_target_rows:
        analog_name = clean_string(row["ANALOG ÖLÇÜM"])
        kind, detail = parse_measure_kind(analog_name)
        metric_key = metric_key_from_kind(kind)
        if not metric_key:
            continue
        candidate = make_scada_candidate(row)
        tm_key = normalize_text(row["TRAFO MERKEZİ"])
        detail_key = normalize_text(detail)

        if "EIH" in detail_key:
            hat = hat_by_name.get(detail_key)
            if hat and metric_key in ("active", "reactive"):
                append_scada_metric(hat["scada"][metric_key], enrich_hat_candidate(candidate, row, hat, tm_by_id))
            continue

        trafo = trafo_by_key.get((tm_key, detail_key))
        if trafo and metric_key in ("active", "reactive", "voltage"):
            append_scada_metric(trafo["scada"][metric_key], candidate)
            continue

        if metric_key == "voltage":
            bara = bara_by_key.get((tm_key, detail_key))
            if bara:
                append_scada_metric(bara["scada"]["voltage"], candidate)

    voltage_exact_source_count = count_metric_presence(bara_nodes, ["scada", "voltage"])
    voltage_overlay_matched = 0
    voltage_alias_matched = 0
    voltage_alias_ambiguous = 0

    chatgpt_voltage_rows = load_table_rows(MATCH_TABLE_FILE, "Gerilim") if MATCH_TABLE_FILE.exists() else []
    if chatgpt_voltage_rows:
        bara_by_id = {bara["id"]: bara for bara in bara_nodes}
        for row in chatgpt_voltage_rows:
            bara_id = as_int_string(get_by_normalized_key(row, "Bara ID"))
            bara = bara_by_id.get(bara_id)
            if not bara:
                continue
            before = bool(bara["scada"]["voltage"]["ids"])
            added = append_scada_metric_once(
                bara["scada"]["voltage"],
                make_voltage_overlay_candidate(row, "chatgpt-voltage-overlay"),
            )
            if added and not before:
                voltage_overlay_matched += 1

    used_voltage_ids = {
        measurement_id
        for bara in bara_nodes
        for measurement_id in bara["scada"]["voltage"]["ids"]
    }
    voltage_alias_index: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in scada_target_rows:
        analog_name = clean_string(row["ANALOG ÖLÇÜM"])
        kind, detail = parse_measure_kind(analog_name)
        if metric_key_from_kind(kind) != "voltage":
            continue
        candidate = make_scada_candidate(row)
        formula_kv_values = [
            part.get("kv")
            for part in candidate.get("formulaParts", [])
            if part.get("parsed")
        ]
        kv_values = {voltage_bucket(detail), *(voltage_bucket(value) for value in formula_kv_values)}
        aliases = set(bara_alias_variants(detail))
        for part in candidate.get("formulaParts", []):
            if part.get("parsed"):
                aliases.update(bara_alias_variants(part.get("targetCode")))
        for kv_value in {value for value in kv_values if value}:
            for alias in aliases:
                voltage_alias_index[(normalize_text(row["TRAFO MERKEZİ"]), kv_value, alias)].append(candidate)

    for bara in bara_nodes:
        if bara["scada"]["voltage"]["ids"]:
            continue
        candidates_by_id: dict[str, dict[str, Any]] = {}
        for alias in bara_alias_variants(bara["name"]):
            for candidate in voltage_alias_index.get((normalize_text(bara["tmName"]), voltage_bucket(bara["gerilimKv"]), alias), []):
                measurement_id = clean_string(candidate.get("measurementId"))
                if not measurement_id or measurement_id in used_voltage_ids:
                    continue
                candidates_by_id[measurement_id] = {**candidate, "source": "voltage-alias-fallback"}
        if len(candidates_by_id) == 1:
            selected = next(iter(candidates_by_id.values()))
            if append_scada_metric_once(bara["scada"]["voltage"], selected):
                used_voltage_ids.add(clean_string(selected.get("measurementId")))
                voltage_alias_matched += 1
        elif len(candidates_by_id) > 1:
            bara["scada"]["voltage"]["ambiguousConfig"] = True
            bara["scada"]["voltage"]["ambiguousConfigIds"] = sorted(candidates_by_id.keys())
            voltage_alias_ambiguous += 1

    for tm in tm_points:
        tm["childTrafoIds"] = sorted(set(tm["childTrafoIds"]), key=sort_key)
        tm["childBaraIds"] = sorted(set(tm["childBaraIds"]), key=sort_key)
        tm["childHatIds"] = sorted(set(tm["childHatIds"]), key=sort_key)

    tm_points.sort(key=lambda item: sort_key(item["id"]))
    hat_lines.sort(key=lambda item: sort_key(item["id"]))
    trafos.sort(key=lambda item: sort_key(item["id"]))
    bara_nodes.sort(key=lambda item: sort_key(item["id"]))

    ytm_names = sorted({tm["ytm"] for tm in tm_points if tm["ytm"]}, key=normalize_text)
    hierarchy = build_hierarchy(tm_points, hat_lines, trafos, bara_nodes)

    hat_scada_rows = [
        metric_row
        for hat in hat_lines
        for metric_key in ("active", "reactive")
        for metric_row in hat["scada"][metric_key]["rows"]
    ]

    validation = {
        "kmlTmCount": len(kml_features["tm"]),
        "kmlHatCount": len(kml_features["hat"]),
        "excelTmCount": len(tm_rows),
        "excelHatCount": len(hat_rows),
        "excelTrafoCount": len(trafo_rows),
        "excelBaraCount": len(bara_rows),
        "tmIdMatches": len(tm_points),
        "hatIdMatches": len(hat_lines),
        "trafoParentMatches": len(trafos),
        "baraParentMatches": len(bara_nodes),
        "bara154400Count": bara_154_400_count,
        "voltageExactSourceMatched": voltage_exact_source_count,
        "voltageOverlayMatched": voltage_overlay_matched,
        "voltageAliasMatched": voltage_alias_matched,
        "voltageAliasAmbiguous": voltage_alias_ambiguous,
        "terminalSideResolved": sum(1 for row in hat_scada_rows if row.get("terminalSide") in {"start", "end"}),
        "terminalSideUnknown": sum(1 for row in hat_scada_rows if row.get("terminalSide") == "unknown"),
        "polarizationMismatch": sum(1 for row in hat_scada_rows if row.get("polarizationConsistent") is False),
        "voltageStillMissing": sum(
            1
            for bara in bara_nodes
            if bara["gerilimKv"] in VALID_BARAGE_SCADA_KV and not bara["scada"]["voltage"]["ids"]
        ),
        "scadaCoverage": {
            "hatActive": count_metric_presence(hat_lines, ["scada", "active"]),
            "hatReactive": count_metric_presence(hat_lines, ["scada", "reactive"]),
            "trafoActive": count_metric_presence(trafos, ["scada", "active"]),
            "trafoReactive": count_metric_presence(trafos, ["scada", "reactive"]),
            "baraVoltageAll": count_metric_presence(bara_nodes, ["scada", "voltage"]),
            "baraVoltage154400": sum(
                1
                for bara in bara_nodes
                if bara["gerilimKv"] in VALID_BARAGE_SCADA_KV and bara["scada"]["voltage"]["ids"]
            ),
        },
        "ambiguous": {
            "hatEntities": count_ambiguous_metrics(hat_lines, [["scada", "active"], ["scada", "reactive"]]),
            "trafoEntities": count_ambiguous_metrics(
                trafos,
                [["scada", "active"], ["scada", "reactive"], ["scada", "voltage"]],
            ),
            "baraEntities": count_ambiguous_metrics(bara_nodes, [["scada", "voltage"]]),
        },
        "missingTmExcelRows": missing_tm_excel_rows,
    }

    model = {
        "meta": {
            "schemaVersion": SCHEMA_VERSION,
            "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "buildTool": "python3 + openpyxl + xml.etree.ElementTree",
            "inputFiles": [
                str(KML_FILE.relative_to(ROOT)),
                str(TM_FILE.relative_to(ROOT)),
                str(HAT_FILE.relative_to(ROOT)),
                str(TRAFO_FILE.relative_to(ROOT)),
                str(BARA_FILE.relative_to(ROOT)),
                str(SCADA_FILE.relative_to(ROOT)),
                str(MATCH_TABLE_FILE.relative_to(ROOT)) if MATCH_TABLE_FILE.exists() else "",
            ],
            "validation": validation,
        },
        "ytmNames": ytm_names,
        "defaultYtm": DEFAULT_YTM if DEFAULT_YTM in ytm_names else (ytm_names[0] if ytm_names else ""),
        "tmPoints": tm_points,
        "hatLines": hat_lines,
        "trafos": trafos,
        "baraNodes": bara_nodes,
        "hierarchy": hierarchy,
    }

    OUTPUT_JSON.write_text(json.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_REPORT.write_text(build_validation_report(model), encoding="utf-8")

    print(f"Wrote {OUTPUT_JSON.relative_to(ROOT)}")
    print(f"Wrote {OUTPUT_REPORT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
